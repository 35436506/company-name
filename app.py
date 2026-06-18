"""
Company Name Fuzzy Matching — Streamlit App
=============================================
Upload either 1 Excel file (multiple tabs) OR 2 separate Excel files.
The user tells the app which tab/file is the **Reference list** (the
standard/master list to match against) and which is the **List to map**
(the list whose company names should be matched against the reference).

Sample data is embedded directly in this file as base64-encoded xlsx
bytes (see SAMPLE DATA section below), so the app is fully self-contained
on Streamlit Cloud's free tier — no external file storage needed. Users
can either download a sample file to inspect/re-upload, or click
"▶ Run this sample" to load it straight into the matcher.

Output: a preview table in the app + a downloadable Excel file with 3 tabs:
  - Result: Reference column, best-match column, match score
  - Source - <Reference list name>: full original data
  - Source - <List to map name>: full original data
"""

import base64
import io
import re
import unicodedata

import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from rapidfuzz import fuzz, process

# ============================================================================
# PAGE CONFIG
# ============================================================================
st.set_page_config(page_title="Company Name Matcher", page_icon="🔎", layout="wide")


# ============================================================================
# SAMPLE DATA (base64-encoded xlsx bytes, embedded directly in this file so
# no external storage is needed — works out of the box on Streamlit Cloud).
# Three samples:
#   1. One workbook with two tabs ("Reference List" + "List To Map")
#   2. A standalone "Reference List" workbook
#   3. A standalone "List To Map" workbook (paired with #2)
# ============================================================================
SAMPLE_1FILE_2TABS_B64 = (
    "UEsDBBQAAAAIADFC0lxGx01IlQAAAM0AAAAQAAAAZG9jUHJvcHMvYXBwLnhtbE3PTQvCMAwG4L9SdreZih6kDkQ9ip68zy51"
    "hbYpbYT67+0EP255ecgboi6JIia2mEXxLuRtMzLHDUDWI/o+y8qhiqHke64x3YGMsRoPpB8eA8OibdeAhTEMOMzit7Dp1C5G"
    "Z3XPlkJ3sjpRJsPiWDQ6sScfq9wcChDneiU+ixNLOZcrBf+LU8sVU57mym/8ZAW/B7oXUEsDBBQAAAAIADFC0lwjwS1w7wAA"
    "ACsCAAARAAAAZG9jUHJvcHMvY29yZS54bWzNks9qwzAMh19l+J7ISUZWTJpLy04bDFbY2M3Yamsa/8HWSPr2S7I2ZWwPsKOl"
    "nz99AjUqCOUjvkQfMJLBdDfYziWhwpodiYIASOqIVqZ8TLixuffRShqf8QBBqpM8IJSc12CRpJYkYQJmYSGyttFKqIiSfLzg"
    "tVrw4TN2M0wrwA4tOkpQ5AWwdpoYzkPXwA0wwQijTd8F1Atxrv6JnTvALskhmSXV933eV3Nu3KGA9+en13ndzLhE0ikcfyUj"
    "6Bxwza6T36rNdvfI2pKXdcbrrFjt+EoUD6K6/5hcf/jdhK3XZm/+sfFVsG3g1120X1BLAwQUAAAACAAxQtJcmVycIxAGAACc"
    "JwAAEwAAAHhsL3RoZW1lL3RoZW1lMS54bWztWltz2jgUfu+v0Hhn9m0LxjaBtrQTc2l227SZhO1OH4URWI1seWSRhH+/RzYQ"
    "y5YN7ZJNups8BCzp+85FR+foOHnz7i5i6IaIlPJ4YNkv29a7ty/e4FcyJBFBMBmnr/DACqVMXrVaaQDDOH3JExLD3IKLCEt4"
    "FMvWXOBbGi8j1uq0291WhGlsoRhHZGB9XixoQNBUUVpvXyC05R8z+BXLVI1lowETV0EmuYi08vlsxfza3j5lz+k6HTKBbjAb"
    "WCB/zm+n5E5aiOFUwsTAamc/VmvH0dJIgILJfZQFukn2o9MVCDINOzqdWM52fPbE7Z+Mytp0NG0a4OPxeDi2y9KLcBwE4FG7"
    "nsKd9Gy/pEEJtKNp0GTY9tqukaaqjVNP0/d93+ubaJwKjVtP02t33dOOicat0HgNvvFPh8Ouicar0HTraSYn/a5rpOkWaEJG"
    "4+t6EhW15UDTIABYcHbWzNIDll4p+nWUGtkdu91BXPBY7jmJEf7GxQTWadIZljRGcp2QBQ4AN8TRTFB8r0G2iuDCktJckNbP"
    "KbVQGgiayIH1R4Ihxdyv/fWXu8mkM3qdfTrOa5R/aasBp+27m8+T/HPo5J+nk9dNQs5wvCwJ8fsjW2GHJ247E3I6HGdCfM/2"
    "9pGlJTLP7/kK6048Zx9WlrBdz8/knoxyI7vd9lh99k9HbiPXqcCzIteURiRFn8gtuuQROLVJDTITPwidhphqUBwCpAkxlqGG"
    "+LTGrBHgE323vgjI342I96tvmj1XoVhJ2oT4EEYa4pxz5nPRbPsHpUbR9lW83KOXWBUBlxjfNKo1LMXWeJXA8a2cPB0TEs2U"
    "CwZBhpckJhKpOX5NSBP+K6Xa/pzTQPCULyT6SpGPabMjp3QmzegzGsFGrxt1h2jSPHr+BfmcNQockRsdAmcbs0YhhGm78B6v"
    "JI6arcIRK0I+Yhk2GnK1FoG2camEYFoSxtF4TtK0EfxZrDWTPmDI7M2Rdc7WkQ4Rkl43Qj5izouQEb8ehjhKmu2icVgE/Z5e"
    "w0nB6ILLZv24fobVM2wsjvdH1BdK5A8mpz/pMjQHo5pZCb2EVmqfqoc0PqgeMgoF8bkePuV6eAo3lsa8UK6CewH/0do3wqv4"
    "gsA5fy59z6XvufQ9odK3NyN9Z8HTi1veRm5bxPuuMdrXNC4oY1dyzcjHVK+TKdg5n8Ds/Wg+nvHt+tkkhK+aWS0jFpBLgbNB"
    "JLj8i8rwKsQJ6GRbJQnLVNNlN4oSnkIbbulT9UqV1+WvuSi4PFvk6a+hdD4sz/k8X+e0zQszQ7dyS+q2lL61JjhK9LHMcE4e"
    "yww7ZzySHbZ3oB01+/ZdduQjpTBTl0O4GkK+A226ndw6OJ6YkbkK01KQb8P56cV4GuI52QS5fZhXbefY0dH758FRsKPvPJYd"
    "x4jyoiHuoYaYz8NDh3l7X5hnlcZQNBRtbKwkLEa3YLjX8SwU4GRgLaAHg69RAvJSVWAxW8YDK5CifEyMRehw55dcX+PRkuPb"
    "pmW1bq8pdxltIlI5wmmYE2eryt5lscFVHc9VW/Kwvmo9tBVOz/5ZrcifDBFOFgsSSGOUF6ZKovMZU77nK0nEVTi/RTO2EpcY"
    "vOPmx3FOU7gSdrYPAjK5uzmpemUxZ6by3y0MCSxbiFkS4k1d7dXnm5yueiJ2+pd3wWDy/XDJRw/lO+df9F1Drn723eP6bpM7"
    "SEycecURAXRFAiOVHAYWFzLkUO6SkAYTAc2UyUTwAoJkphyAmPoLvfIMuSkVzq0+OX9FLIOGTl7SJRIUirAMBSEXcuPv75Nq"
    "d4zX+iyBbYRUMmTVF8pDicE9M3JD2FQl867aJguF2+JUzbsaviZgS8N6bp0tJ//bXtQ9tBc9RvOjmeAes4dzm3q4wkWs/1jW"
    "Hvky3zlw2zreA17mEyxDpH7BfYqKgBGrYr66r0/5JZw7tHvxgSCb/NbbpPbd4Ax81KtapWQrET9LB3wfkgZjjFv0NF+PFGKt"
    "prGtxtoxDHmAWPMMoWY434dFmhoz1YusOY0Kb0HVQOU/29QNaPYNNByRBV4xmbY2o+ROCjzc/u8NsMLEjuHti78BUEsDBBQA"
    "AAAIADFC0lwawfSlwQQAAJEVAAAYAAAAeGwvd29ya3NoZWV0cy9zaGVldDEueG1spZhdc+I2FIb/isY7k7vG2OQDNsBMQpJN"
    "WpLSsG1neqfYx6CJLLmyXJb++h5ZrE2nQmKmN/jzfdFjyTqvNdlK9VFvADT5VnJRT6ON1tXnOK6zDZS0PpcVCLxSSFVSjYdq"
    "HdeVApq3opLH6WBwFZeUiWg2ac8t1WwiG82ZgKUidVOWVO3ugMvtNEqi7yfe2HqjzYl4NqnoGlagf62WCo/iziVnJYiaSUEU"
    "FNPoNvk8T4dG0N7xG4NtfbBPDMq7lB/m4DmfRoPIWAsgu1XFWftnRMtqAYWeA+domEaEZpr9BUu8bRq9S61laa5jMzXVeKpQ"
    "8m8Q7X8CB7wXG1P952Zrsjc1jH/uGxx1PKZRh/vfW/7YPlh8UO+0hrnkv7Ncb6bRKCI5FLTh+k1un2D/sC6NXyZ53f6Srb33"
    "YhyRrKmxNXsxtqBkwm7pt/1DPhAkoyOCdC9ITxUM94K2V2Lbshbrnmo6myi5JcrcjW5mp302th+mERNmiKy0wqsMdXo2lyU+"
    "2x15pSVMYo2O5nyc7dV3fvUbFKBAZECe7x3quV+9wv5u6n/rYmx/B5F2EGlrlB4xus1KIF8VzZlYkwUrmYbcBeN3eXt4/GEw"
    "SFwcob83Q9HDMew4hl6jO9CUrCRvzJivyUI7KfweliJ1UfiFQYqLjuLCazSnxdmndDi8Ia+SKbK6dUH4LSzE0AXhFwYhLjuI"
    "S6/RPXDsiyfJzZiqyXIxJzqmNbHnvyjZVC4sv6nFunBh+YVBrKsO68pr9FDVjON8+ixynFsUgxp3s3MXit/Ioly6UPzCIMp1"
    "h3LtNXqkPEOSM1pWN/jSmPfl+Jvvt7IwVy4YvzAIM+pgRl6jL4oKbDtZyDWrNcsQZrF0gfhtLMi1C8QvDIKMO5Cx1+iJqnep"
    "SBsOnqQGfnQa8/tYkpGLxC8MkiSDvjQOvFbPSoqM05ysIGsU07vDufn4WAu4WrKxs1z6lWG0g6qfeK1+bCpsviIvkLP9dGbm"
    "OCeO36nFSQZOHL8yjNPX/8RfgX/CSbpgmIIUecMayvixQRfwsTDOCBBQhmH6EJD4C/HCJM+NbGogcxxqmEfbWHOkd05IA4kz"
    "DQSUYaA+DyT+ovwCiuEwE/j6FHpLFZia44Q5IRUkzlQQUIZh+lyQ+Evxq1R6s2UiJw9ijdeQzR86A34WypkJAsowVJ8KEn9R"
    "/pl+YBetgSwVfnwqbXIBvkE28nQXH9oPtNoJeUJaSJxpIaAMQ/Z5IfEX6yUTgmYcyNmnZHA1uDGb0X5z2W6u08Oj0bjdjMc3"
    "R2eTE2JF4owVAWWYug8Wib+y/9JQoZuS3ArKdzZZeMbqCfEiccaLgDIM1AeMJJAM0EjVLAfyKGVuxykT5DbHr2PMToqaAu2E"
    "OyFxJM7EEVAG4dI+cqT+Er9qSuwc8gSU642/LgecLI4zZgSUYZw+ZqT+Ev8VJw4QmrxQ0RQ00007Wc6lk+eEnJE6c0ZAGeY5"
    "WGfwV/k7jOk5ecVsu9JUaeycIzNDwMfCuJca/mfOSPuckfor/B9QbXbqMGMc/awNOFkcZ8oIKI/jxAdrWmax8oUqLLc14VCg"
    "0+D8Goujsmt09kDLql1qsouE7e4GaA7K3IDXC4mfI/sDs3LWrcLO/gFQSwMEFAAAAAgAMULSXFRzwi0ZBAAA6A8AABgAAAB4"
    "bC93b3Jrc2hlZXRzL3NoZWV0Mi54bWyNl11zmzgUhv+KhvstFonz0bE9k7hNnV0n9dpt91qBg62JkFhJrOv99XsEBLxTfOKb"
    "hA/p9YMkHnQme2Nf3Q7As5+F0m4a7bwvP8axS3dQCPfBlKDxTm5sITye2m3sSgsiqzsVKk5Go6u4EFJHs0l9bWVnE1N5JTWs"
    "LHNVUQh7uAdl9tOIR28X1nK78+FCPJuUYgsb8N/LlcWzuEvJZAHaSaOZhXwa3fGP98lF6FC3+CFh746OWXiUF2New8ljNo1G"
    "UYjWwA6bUsn6x5g35RJyPwelMDCJmEi9/AdW2GwavRjvTRHuI6YXHi/l1vwLuv5NUIBtEab8pXET0oaGZ/y7BY665wlQx8dv"
    "5A/1wOJAvQgHc6P+kpnfTaObiGWQi0r5tdkvoB2scchLjXL1X7Zv2l6MI5ZWDmnazkhQSN38Fz/bQT7qwG9OdEjaDknN3fxQ"
    "TflJeDGbWLNnNrTGtHBQP2ozrNNI6jDjG2/xrsR+fjY3BQ7VgT2LAiaxx8RwPU7b3vd077s0NZX27LkqXsD+v3+MJB1O0uEk"
    "dWByMrAA9s2KTOotW/psCOmdhPn8Nz4ajQiYiw7mgoy6By/YxqgqrCjHlrKQHgaR6JwWiRNIlx3SJRk1FzmwZyPtEAXdtaVI"
    "CIpxRzEmoz6BwpFZGBWmyQ2h0P1blAsC5apDuSKjPpdOKnzfH3WGL4uV4PAwNbY0VpyYLDqwZbsk2K47tmsy6kGoFNGEznAV"
    "hQU0vJ7pkJZnTPDcdDw3ZNQXKzSuYLY0W+m8TAdnjo5oaa4ImtuO5paMWgj7YirL6s/CwnhQ5CtGh7Vc1wQXH/VKHJFhj9bo"
    "VAmcNUgrK/2hl8CgIum0Fu2GQjuyNSfDfq9KHCDLniCTgn2xpioHkeiUFumWQuqNzWnh/oEOyCV+gSxjazSmVIyax3fSGjRO"
    "6Zv3/ua0eJfhk7wzlQM2x8nDDzWyDkKdI3BOCZz3Bue0h5/ASpw8jYsq93thgRytc5zOKafzXuqctvKzsX63l2irz3oLGpDz"
    "9Ef4nawWjDI87xXPaSV/Fa84ZFtgK4s7XetR8oNI54idU2Lnvdk5beWV1FqkCtiD1EK7w8lhOkfunJI77+3OaTf/WQntq4Ld"
    "aaEOp/T+TkYLRPmd94LntJTXuN+2TmY4SsZkwzjnaJ1TWk96rSe0iBcglN810mQbrHCkZ6vlfHCLeY7SOaX0pFd6Qsv4G65t"
    "wD30k9BVjkVKVb927b58EO4cuXNK7snRdpzW8XdtQYWNFFvjNsYU7L5y2MSd3My8k9fAJZTek17vCa3ljcFi4asPX563Mgb3"
    "foNU5/g9GfR7fFRbhRr4SditxM2cwqISy9cP1yhC25R+zQlWr3XB1NSe9eEOS3GwoQHezw3uddqTUMF1xf3sP1BLAwQUAAAA"
    "CAAxQtJch72R07gCAAC9CwAADQAAAHhsL3N0eWxlcy54bWzdVtuOmzAQ/RXEB5QkpDRUEGmLtFKltlpp96GvTjBgyRdqzCrZ"
    "r68HE8jFs7ut+lRQhD3H58zFY0jWmSOnjw2lJjgILrs8bIxpP0dRt2+oIN0H1VJpkUppQYyd6jrqWk1J2QFJ8Gi1WCSRIEyG"
    "20z24l6YLtirXpo8XITRNquUnC1x6Ax2KRE0eCY8DwvC2U6zYS0RjB+deQWGveJKB8aGQvNwCZbuxcFLN4MoRx3BpNJgjJyH"
    "az93mhEO+G5UmB3oemejXdwP162XtwSn1Yuz1cOjsyzG+WX+1rDNWmIM1fLeTgbOYLyBgnH8dGxtAWpNjsvVx/DdhE5xVoLL"
    "ujjPc73+tCrWg8wZdRIdHjbyndIl1VPsy/Bk2macVsbSNasbeBrVQl2VMUrYQclIrSQZEjsxxoGV3VPOH6HtflYX2ocqcP3z"
    "tYTWCaB+p6ENaBw6GTcB/XM1p30mG/+VbNCyZ2W+9DYbOcx/9crQB00rdhjmh2ryj6kvZ/XVlTppW36846yWgrrc3+1wm5ET"
    "L2iUZi/WGzTe3hqo6/1DhQe1+icpR2ORz3byYh8nawBHJQ9/wKuDzxLBrmfcMDnOGlaWVN5sp5U3ZGffTRf6dn1JK9Jz8zSB"
    "eTiPv9OS9SKdVj1AWuOqefwN2naZTOff+mKypAdaFuPUHpSLE+MuIFwj8zvjFsE4DvMjgGF+sAgwjmNhfv6nfDZoPg7DYtt4"
    "kQ3K2aAcx/IhxXBjfvyc1F7+TNM0jpMEq2hReCMosLolCfz8alhswMD8gKc/qzW+23iHvN4H2J6+1iFYpngnYpnitQbEXzdg"
    "pKl/tzE/wMB2Aesd8O/3Az3l58Qx7CoWG3aCcSRNMQR60d+jSYJUJ4Hbvz/YKYnjNPUjgPkjiGMMgdOII1gEEAOGxPHwHbz6"
    "HkWn71Q0/2Hf/gZQSwMEFAAAAAgAMULSXJeKuxzAAAAAEwIAAAsAAABfcmVscy8ucmVsc52SuW7DMAxAf8XQnjAH0CGIM2Xx"
    "FgT5AVaiD9gSBYpFnb+v2qVxkAsZeT08EtweaUDtOKS2i6kY/RBSaVrVuAFItiWPac6RQq7ULB41h9JARNtjQ7BaLD5ALhlm"
    "t71kFqdzpFeIXNedpT3bL09Bb4CvOkxxQmlISzMO8M3SfzL38ww1ReVKI5VbGnjT5f524EnRoSJYFppFydOiHaV/Hcf2kNPp"
    "r2MitHpb6PlxaFQKjtxjJYxxYrT+NYLJD+x+AFBLAwQUAAAACAAxQtJcMvG0wFABAAC8AgAADwAAAHhsL3dvcmtib29rLnht"
    "bLWSYUvDMBCG/0rID7Dd0IFj9YtDHUwd29j3rL2ux5JcuVw33a83bSkOBPGLn5J7L7x57k1mZ+LjnuioPpz1IdOVSD1NkpBX"
    "4Ey4oRp87JTEzkgs+ZCEmsEUoQIQZ5Nxmk4SZ9Drh9ngteLkuiCBXJB8FFthh3AO3/22VCcMuEeL8pnpbm9BK4ceHV6gyHSq"
    "Vajo/EKMF/Ji7CZnsjbTo76xAxbMf8ibFnJr9qFTxOzXJoJkepJGwxI5SHei8zeR8QTxcF81Qk9oBXhuBJ6Zmhr9obWJUyRX"
    "Y3Q5DGsf4pT/EiOVJeYwp7xx4KXPkcG2gD5UWAetvHGQ6TWUwOBzUEsM0g4Wb1oU/ZAS6a4i4ynGBi+KjvP/mFoStSX1auor"
    "oPEvQOMuuCGtAkr0ULxFsxD1+HL5ilW7dIONb+9G9/GFGmsfo/bul2SKIfzh4zx8AVBLAwQUAAAACAAxQtJcjfcsWrQAAACJ"
    "AgAAGgAAAHhsL19yZWxzL3dvcmtib29rLnhtbC5yZWxzxZJNCoMwEEavEnKAjtrSRVFX3bgtXiDo+IPRhMyU6u1rdaGBLrqR"
    "rsI3Ie97MIkfqBW3ZqCmtSTGXg+UyIbZ3gCoaLBXdDIWh/mmMq5XPEdXg1VFp2qEKAiu4PYMmcZ7psgni78QTVW1Bd5N8exx"
    "4C9geBnXUYPIUuTK1ciJhFFvY4LlCE8zWYqsTKTLylDCv4UiTyg6UIh40kibzZq9+vOB9Ty/xa19ievQ38nl4wDez0vfUEsD"
    "BBQAAAAIADFC0lxupyS8HgEAAFcEAAATAAAAW0NvbnRlbnRfVHlwZXNdLnhtbMWUz07DMAzGX6XKdWoyduCA1l2AK+zAC4TW"
    "XaPmn2JvdG+P226TQKNiKhKXRo3t7+f4i7J+O0bArHPWYyEaovigFJYNOI0yRPAcqUNymvg37VTUZat3oFbL5b0qgyfwlFOv"
    "ITbrJ6j13lL23PE2muALkcCiyB7HxJ5VCB2jNaUmjquDr75R8hNBcuWQg42JuOAEoa4S+sjPgFPd6wFSMhVkW53oRTvOUp1V"
    "SEcLKKclrvQY6tqUUIVy77hEYkygK2wAyFk5ii6mycQThvF7N5s/yEwBOXObQkR2LMHtuLMlfXUeWQgSmekjXogsPft80Ltd"
    "QfVLNo/3I6R28APVsMyf8VePL/o39rH6xz7eQ2j/+qr3q3Ta+DNfDe/J5hNQSwECFAMUAAAACAAxQtJcRsdNSJUAAADNAAAA"
    "EAAAAAAAAAAAAAAAgAEAAAAAZG9jUHJvcHMvYXBwLnhtbFBLAQIUAxQAAAAIADFC0lwjwS1w7wAAACsCAAARAAAAAAAAAAAA"
    "AACAAcMAAABkb2NQcm9wcy9jb3JlLnhtbFBLAQIUAxQAAAAIADFC0lyZXJwjEAYAAJwnAAATAAAAAAAAAAAAAACAAeEBAAB4"
    "bC90aGVtZS90aGVtZTEueG1sUEsBAhQDFAAAAAgAMULSXBrB9KXBBAAAkRUAABgAAAAAAAAAAAAAAICBIggAAHhsL3dvcmtz"
    "aGVldHMvc2hlZXQxLnhtbFBLAQIUAxQAAAAIADFC0lxUc8ItGQQAAOgPAAAYAAAAAAAAAAAAAACAgRkNAAB4bC93b3Jrc2hl"
    "ZXRzL3NoZWV0Mi54bWxQSwECFAMUAAAACAAxQtJch72R07gCAAC9CwAADQAAAAAAAAAAAAAAgAFoEQAAeGwvc3R5bGVzLnht"
    "bFBLAQIUAxQAAAAIADFC0lyXirscwAAAABMCAAALAAAAAAAAAAAAAACAAUsUAABfcmVscy8ucmVsc1BLAQIUAxQAAAAIADFC"
    "0lwy8bTAUAEAALwCAAAPAAAAAAAAAAAAAACAATQVAAB4bC93b3JrYm9vay54bWxQSwECFAMUAAAACAAxQtJcjfcsWrQAAACJ"
    "AgAAGgAAAAAAAAAAAAAAgAGxFgAAeGwvX3JlbHMvd29ya2Jvb2sueG1sLnJlbHNQSwECFAMUAAAACAAxQtJcbqckvB4BAABX"
    "BAAAEwAAAAAAAAAAAAAAgAGdFwAAW0NvbnRlbnRfVHlwZXNdLnhtbFBLBQYAAAAACgAKAIQCAADsGAAAAAA="
)

SAMPLE_REFERENCE_LIST_B64 = (
    "UEsDBBQAAAAIADFC0lxGx01IlQAAAM0AAAAQAAAAZG9jUHJvcHMvYXBwLnhtbE3PTQvCMAwG4L9SdreZih6kDkQ9ip68zy51"
    "hbYpbYT67+0EP255ecgboi6JIia2mEXxLuRtMzLHDUDWI/o+y8qhiqHke64x3YGMsRoPpB8eA8OibdeAhTEMOMzit7Dp1C5G"
    "Z3XPlkJ3sjpRJsPiWDQ6sScfq9wcChDneiU+ixNLOZcrBf+LU8sVU57mym/8ZAW/B7oXUEsDBBQAAAAIADFC0lwjwS1w7wAA"
    "ACsCAAARAAAAZG9jUHJvcHMvY29yZS54bWzNks9qwzAMh19l+J7ISUZWTJpLy04bDFbY2M3Yamsa/8HWSPr2S7I2ZWwPsKOl"
    "nz99AjUqCOUjvkQfMJLBdDfYziWhwpodiYIASOqIVqZ8TLixuffRShqf8QBBqpM8IJSc12CRpJYkYQJmYSGyttFKqIiSfLzg"
    "tVrw4TN2M0wrwA4tOkpQ5AWwdpoYzkPXwA0wwQijTd8F1Atxrv6JnTvALskhmSXV933eV3Nu3KGA9+en13ndzLhE0ikcfyUj"
    "6Bxwza6T36rNdvfI2pKXdcbrrFjt+EoUD6K6/5hcf/jdhK3XZm/+sfFVsG3g1120X1BLAwQUAAAACAAxQtJcmVycIxAGAACc"
    "JwAAEwAAAHhsL3RoZW1lL3RoZW1lMS54bWztWltz2jgUfu+v0Hhn9m0LxjaBtrQTc2l227SZhO1OH4URWI1seWSRhH+/RzYQ"
    "y5YN7ZJNups8BCzp+85FR+foOHnz7i5i6IaIlPJ4YNkv29a7ty/e4FcyJBFBMBmnr/DACqVMXrVaaQDDOH3JExLD3IKLCEt4"
    "FMvWXOBbGi8j1uq0291WhGlsoRhHZGB9XixoQNBUUVpvXyC05R8z+BXLVI1lowETV0EmuYi08vlsxfza3j5lz+k6HTKBbjAb"
    "WCB/zm+n5E5aiOFUwsTAamc/VmvH0dJIgILJfZQFukn2o9MVCDINOzqdWM52fPbE7Z+Mytp0NG0a4OPxeDi2y9KLcBwE4FG7"
    "nsKd9Gy/pEEJtKNp0GTY9tqukaaqjVNP0/d93+ubaJwKjVtP02t33dOOicat0HgNvvFPh8Ouicar0HTraSYn/a5rpOkWaEJG"
    "4+t6EhW15UDTIABYcHbWzNIDll4p+nWUGtkdu91BXPBY7jmJEf7GxQTWadIZljRGcp2QBQ4AN8TRTFB8r0G2iuDCktJckNbP"
    "KbVQGgiayIH1R4Ihxdyv/fWXu8mkM3qdfTrOa5R/aasBp+27m8+T/HPo5J+nk9dNQs5wvCwJ8fsjW2GHJ247E3I6HGdCfM/2"
    "9pGlJTLP7/kK6048Zx9WlrBdz8/knoxyI7vd9lh99k9HbiPXqcCzIteURiRFn8gtuuQROLVJDTITPwidhphqUBwCpAkxlqGG"
    "+LTGrBHgE323vgjI342I96tvmj1XoVhJ2oT4EEYa4pxz5nPRbPsHpUbR9lW83KOXWBUBlxjfNKo1LMXWeJXA8a2cPB0TEs2U"
    "CwZBhpckJhKpOX5NSBP+K6Xa/pzTQPCULyT6SpGPabMjp3QmzegzGsFGrxt1h2jSPHr+BfmcNQockRsdAmcbs0YhhGm78B6v"
    "JI6arcIRK0I+Yhk2GnK1FoG2camEYFoSxtF4TtK0EfxZrDWTPmDI7M2Rdc7WkQ4Rkl43Qj5izouQEb8ehjhKmu2icVgE/Z5e"
    "w0nB6ILLZv24fobVM2wsjvdH1BdK5A8mpz/pMjQHo5pZCb2EVmqfqoc0PqgeMgoF8bkePuV6eAo3lsa8UK6CewH/0do3wqv4"
    "gsA5fy59z6XvufQ9odK3NyN9Z8HTi1veRm5bxPuuMdrXNC4oY1dyzcjHVK+TKdg5n8Ds/Wg+nvHt+tkkhK+aWS0jFpBLgbNB"
    "JLj8i8rwKsQJ6GRbJQnLVNNlN4oSnkIbbulT9UqV1+WvuSi4PFvk6a+hdD4sz/k8X+e0zQszQ7dyS+q2lL61JjhK9LHMcE4e"
    "yww7ZzySHbZ3oB01+/ZdduQjpTBTl0O4GkK+A226ndw6OJ6YkbkK01KQb8P56cV4GuI52QS5fZhXbefY0dH758FRsKPvPJYd"
    "x4jyoiHuoYaYz8NDh3l7X5hnlcZQNBRtbKwkLEa3YLjX8SwU4GRgLaAHg69RAvJSVWAxW8YDK5CifEyMRehw55dcX+PRkuPb"
    "pmW1bq8pdxltIlI5wmmYE2eryt5lscFVHc9VW/Kwvmo9tBVOz/5ZrcifDBFOFgsSSGOUF6ZKovMZU77nK0nEVTi/RTO2EpcY"
    "vOPmx3FOU7gSdrYPAjK5uzmpemUxZ6by3y0MCSxbiFkS4k1d7dXnm5yueiJ2+pd3wWDy/XDJRw/lO+df9F1Drn723eP6bpM7"
    "SEycecURAXRFAiOVHAYWFzLkUO6SkAYTAc2UyUTwAoJkphyAmPoLvfIMuSkVzq0+OX9FLIOGTl7SJRIUirAMBSEXcuPv75Nq"
    "d4zX+iyBbYRUMmTVF8pDicE9M3JD2FQl867aJguF2+JUzbsaviZgS8N6bp0tJ//bXtQ9tBc9RvOjmeAes4dzm3q4wkWs/1jW"
    "Hvky3zlw2zreA17mEyxDpH7BfYqKgBGrYr66r0/5JZw7tHvxgSCb/NbbpPbd4Ax81KtapWQrET9LB3wfkgZjjFv0NF+PFGKt"
    "prGtxtoxDHmAWPMMoWY434dFmhoz1YusOY0Kb0HVQOU/29QNaPYNNByRBV4xmbY2o+ROCjzc/u8NsMLEjuHti78BUEsDBBQA"
    "AAAIADFC0lwawfSlwQQAAJEVAAAYAAAAeGwvd29ya3NoZWV0cy9zaGVldDEueG1spZhdc+I2FIb/isY7k7vG2OQDNsBMQpJN"
    "WpLSsG1neqfYx6CJLLmyXJb++h5ZrE2nQmKmN/jzfdFjyTqvNdlK9VFvADT5VnJRT6ON1tXnOK6zDZS0PpcVCLxSSFVSjYdq"
    "HdeVApq3opLH6WBwFZeUiWg2ac8t1WwiG82ZgKUidVOWVO3ugMvtNEqi7yfe2HqjzYl4NqnoGlagf62WCo/iziVnJYiaSUEU"
    "FNPoNvk8T4dG0N7xG4NtfbBPDMq7lB/m4DmfRoPIWAsgu1XFWftnRMtqAYWeA+domEaEZpr9BUu8bRq9S61laa5jMzXVeKpQ"
    "8m8Q7X8CB7wXG1P952Zrsjc1jH/uGxx1PKZRh/vfW/7YPlh8UO+0hrnkv7Ncb6bRKCI5FLTh+k1un2D/sC6NXyZ53f6Srb33"
    "YhyRrKmxNXsxtqBkwm7pt/1DPhAkoyOCdC9ITxUM94K2V2Lbshbrnmo6myi5JcrcjW5mp302th+mERNmiKy0wqsMdXo2lyU+"
    "2x15pSVMYo2O5nyc7dV3fvUbFKBAZECe7x3quV+9wv5u6n/rYmx/B5F2EGlrlB4xus1KIF8VzZlYkwUrmYbcBeN3eXt4/GEw"
    "SFwcob83Q9HDMew4hl6jO9CUrCRvzJivyUI7KfweliJ1UfiFQYqLjuLCazSnxdmndDi8Ia+SKbK6dUH4LSzE0AXhFwYhLjuI"
    "S6/RPXDsiyfJzZiqyXIxJzqmNbHnvyjZVC4sv6nFunBh+YVBrKsO68pr9FDVjON8+ixynFsUgxp3s3MXit/Ioly6UPzCIMp1"
    "h3LtNXqkPEOSM1pWN/jSmPfl+Jvvt7IwVy4YvzAIM+pgRl6jL4oKbDtZyDWrNcsQZrF0gfhtLMi1C8QvDIKMO5Cx1+iJqnep"
    "SBsOnqQGfnQa8/tYkpGLxC8MkiSDvjQOvFbPSoqM05ysIGsU07vDufn4WAu4WrKxs1z6lWG0g6qfeK1+bCpsviIvkLP9dGbm"
    "OCeO36nFSQZOHL8yjNPX/8RfgX/CSbpgmIIUecMayvixQRfwsTDOCBBQhmH6EJD4C/HCJM+NbGogcxxqmEfbWHOkd05IA4kz"
    "DQSUYaA+DyT+ovwCiuEwE/j6FHpLFZia44Q5IRUkzlQQUIZh+lyQ+Evxq1R6s2UiJw9ijdeQzR86A34WypkJAsowVJ8KEn9R"
    "/pl+YBetgSwVfnwqbXIBvkE28nQXH9oPtNoJeUJaSJxpIaAMQ/Z5IfEX6yUTgmYcyNmnZHA1uDGb0X5z2W6u08Oj0bjdjMc3"
    "R2eTE2JF4owVAWWYug8Wib+y/9JQoZuS3ArKdzZZeMbqCfEiccaLgDIM1AeMJJAM0EjVLAfyKGVuxykT5DbHr2PMToqaAu2E"
    "OyFxJM7EEVAG4dI+cqT+Er9qSuwc8gSU642/LgecLI4zZgSUYZw+ZqT+Ev8VJw4QmrxQ0RQ00007Wc6lk+eEnJE6c0ZAGeY5"
    "WGfwV/k7jOk5ecVsu9JUaeycIzNDwMfCuJca/mfOSPuckfor/B9QbXbqMGMc/awNOFkcZ8oIKI/jxAdrWmax8oUqLLc14VCg"
    "0+D8Goujsmt09kDLql1qsouE7e4GaA7K3IDXC4mfI/sDs3LWrcLO/gFQSwMEFAAAAAgAMULSXIe9kdO4AgAAvQsAAA0AAAB4"
    "bC9zdHlsZXMueG1s3VbbjpswEP0VxAeUJKQ0VBBpi7RSpbZaafehr04wYMkXaswq2a+vBxPIxbO7rfpUUIQ9x+fMxWNI1pkj"
    "p48NpSY4CC67PGyMaT9HUbdvqCDdB9VSaZFKaUGMneo66lpNSdkBSfBotVgkkSBMhttM9uJemC7Yq16aPFyE0TarlJwtcegM"
    "dikRNHgmPA8LwtlOs2EtEYwfnXkFhr3iSgfGhkLzcAmW7sXBSzeDKEcdwaTSYIych2s/d5oRDvhuVJgd6Hpno13cD9etl7cE"
    "p9WLs9XDo7Msxvll/tawzVpiDNXy3k4GzmC8gYJx/HRsbQFqTY7L1cfw3YROcVaCy7o4z3O9/rQq1oPMGXUSHR428p3SJdVT"
    "7MvwZNpmnFbG0jWrG3ga1UJdlTFK2EHJSK0kGRI7McaBld1Tzh+h7X5WF9qHKnD987WE1gmgfqehDWgcOhk3Af1zNad9Jhv/"
    "lWzQsmdlvvQ2GznMf/XK0AdNK3YY5odq8o+pL2f11ZU6aVt+vOOsloK63N/tcJuREy9olGYv1hs03t4aqOv9Q4UHtfonKUdj"
    "kc928mIfJ2sARyUPf8Crg88Swa5n3DA5zhpWllTebKeVN2Rn300X+nZ9SSvSc/M0gXk4j7/TkvUinVY9QFrjqnn8Ddp2mUzn"
    "3/pisqQHWhbj1B6UixPjLiBcI/M74xbBOA7zI4BhfrAIMI5jYX7+p3w2aD4Ow2LbeJENytmgHMfyIcVwY378nNRe/kzTNI6T"
    "BKtoUXgjKLC6JQn8/GpYbMDA/ICnP6s1vtt4h7zeB9ievtYhWKZ4J2KZ4rUGxF83YKSpf7cxP8DAdgHrHfDv9wM95efEMewq"
    "Fht2gnEkTTEEetHfo0mCVCeB278/2CmJ4zT1I4D5I4hjDIHTiCNYBBADhsTx8B28+h5Fp+9UNP9h3/4GUEsDBBQAAAAIADFC"
    "0lyXirscwAAAABMCAAALAAAAX3JlbHMvLnJlbHOdkrluwzAMQH/F0J4wB9AhiDNl8RYE+QFWog/YEgWKRZ2/r9qlcZALGXk9"
    "PBLcHmlA7TiktoupGP0QUmla1bgBSLYlj2nOkUKu1CweNYfSQETbY0OwWiw+QC4ZZre9ZBanc6RXiFzXnaU92y9PQW+ArzpM"
    "cUJpSEszDvDN0n8y9/MMNUXlSiOVWxp40+X+duBJ0aEiWBaaRcnToh2lfx3H9pDT6a9jIrR6W+j5cWhUCo7cYyWMcWK0/jWC"
    "yQ/sfgBQSwMEFAAAAAgAMULSXFuRv2s5AQAAKwIAAA8AAAB4bC93b3JrYm9vay54bWyNUdFuwjAM/JUqH7CWaUMaorwMbUNC"
    "G4KJ99C61CKJK8eFja+f26oa0l72lNzZutxd5hfi04HolHx5F2JuapFmlqaxqMHbeEcNBJ1UxN6KQj6msWGwZawBxLv0Psum"
    "qbcYzGI+am04vQUkUAhSULIj9giX+DvvYHLGiAd0KN+56e8OTOIxoMcrlLnJTBJrurwR45WCWLcrmJzLzWQY7IEFiz/0rjP5"
    "aQ+xZ8QetlaN5GaaqWCFHKXf6PWtejyDLg+oFXpBJ8BLK/DK1DYYjp2MpkhvYvQ9jOdQ4oz/UyNVFRawpKL1EGTokcF1BkOs"
    "sYkmCdZDbrZQAUMoIFljlC6YvrQqh5Ci7m4q4xnqgFfl4HM0V0KFAcp31YvKa1HFhpPu6HXuHx4nT1pI69yzch9hTbYcs47/"
    "tPgBUEsDBBQAAAAIADFC0lwkHpuirQAAAPgBAAAaAAAAeGwvX3JlbHMvd29ya2Jvb2sueG1sLnJlbHO1kT0OgzAMha8S5QA1"
    "UKlDBUxdWCsuEAXzIxISxa4Kty+FAZA6dGGyni1/78lOn2gUd26gtvMkRmsGymTL7O8ApFu0ii7O4zBPahes4lmGBrzSvWoQ"
    "kii6QdgzZJ7umaKcPP5DdHXdaXw4/bI48A8wvF3oqUVkKUoVGuRMwmi2NsFS4stMlqKoMhmKKpZwWiDiySBtaVZ9sE9OtOd5"
    "Fzf3Ra7N4wmu3wxweHT+AVBLAwQUAAAACAAxQtJcZZB5khkBAADPAwAAEwAAAFtDb250ZW50X1R5cGVzXS54bWytk01OwzAQ"
    "ha8SZVslLixYoKYbYAtdcAFjTxqr/pNnWtLbM07aSqASFYVNrHjevM+el6zejxGw6J312JQdUXwUAlUHTmIdIniutCE5Sfya"
    "tiJKtZNbEPfL5YNQwRN4qih7lOvVM7Ryb6l46XkbTfBNmcBiWTyNwsxqShmjNUoS18XB6x+U6kSouXPQYGciLlhQiquEXPkd"
    "cOp7O0BKRkOxkYlepWOV6K1AOlrAetriyhlD2xoFOqi945YaYwKpsQMgZ+vRdDFNJp4wjM+72fzBZgrIyk0KETmxBH/HnSPJ"
    "3VVkI0hkpq94IbL17PtBTluDvpHN4/0MaTfkgWJY5s/4e8YX/xvO8RHC7r8/sbzWThp/5ovhP15/AVBLAQIUAxQAAAAIADFC"
    "0lxGx01IlQAAAM0AAAAQAAAAAAAAAAAAAACAAQAAAABkb2NQcm9wcy9hcHAueG1sUEsBAhQDFAAAAAgAMULSXCPBLXDvAAAA"
    "KwIAABEAAAAAAAAAAAAAAIABwwAAAGRvY1Byb3BzL2NvcmUueG1sUEsBAhQDFAAAAAgAMULSXJlcnCMQBgAAnCcAABMAAAAA"
    "AAAAAAAAAIAB4QEAAHhsL3RoZW1lL3RoZW1lMS54bWxQSwECFAMUAAAACAAxQtJcGsH0pcEEAACRFQAAGAAAAAAAAAAAAAAA"
    "gIEiCAAAeGwvd29ya3NoZWV0cy9zaGVldDEueG1sUEsBAhQDFAAAAAgAMULSXIe9kdO4AgAAvQsAAA0AAAAAAAAAAAAAAIAB"
    "GQ0AAHhsL3N0eWxlcy54bWxQSwECFAMUAAAACAAxQtJcl4q7HMAAAAATAgAACwAAAAAAAAAAAAAAgAH8DwAAX3JlbHMvLnJl"
    "bHNQSwECFAMUAAAACAAxQtJcW5G/azkBAAArAgAADwAAAAAAAAAAAAAAgAHlEAAAeGwvd29ya2Jvb2sueG1sUEsBAhQDFAAA"
    "AAgAMULSXCQem6KtAAAA+AEAABoAAAAAAAAAAAAAAIABSxIAAHhsL19yZWxzL3dvcmtib29rLnhtbC5yZWxzUEsBAhQDFAAA"
    "AAgAMULSXGWQeZIZAQAAzwMAABMAAAAAAAAAAAAAAIABMBMAAFtDb250ZW50X1R5cGVzXS54bWxQSwUGAAAAAAkACQA+AgAA"
    "ehQAAAAA"
)

SAMPLE_LIST_TO_MAP_B64 = (
    "UEsDBBQAAAAIADFC0lxGx01IlQAAAM0AAAAQAAAAZG9jUHJvcHMvYXBwLnhtbE3PTQvCMAwG4L9SdreZih6kDkQ9ip68zy51"
    "hbYpbYT67+0EP255ecgboi6JIia2mEXxLuRtMzLHDUDWI/o+y8qhiqHke64x3YGMsRoPpB8eA8OibdeAhTEMOMzit7Dp1C5G"
    "Z3XPlkJ3sjpRJsPiWDQ6sScfq9wcChDneiU+ixNLOZcrBf+LU8sVU57mym/8ZAW/B7oXUEsDBBQAAAAIADFC0lwjwS1w7wAA"
    "ACsCAAARAAAAZG9jUHJvcHMvY29yZS54bWzNks9qwzAMh19l+J7ISUZWTJpLy04bDFbY2M3Yamsa/8HWSPr2S7I2ZWwPsKOl"
    "nz99AjUqCOUjvkQfMJLBdDfYziWhwpodiYIASOqIVqZ8TLixuffRShqf8QBBqpM8IJSc12CRpJYkYQJmYSGyttFKqIiSfLzg"
    "tVrw4TN2M0wrwA4tOkpQ5AWwdpoYzkPXwA0wwQijTd8F1Atxrv6JnTvALskhmSXV933eV3Nu3KGA9+en13ndzLhE0ikcfyUj"
    "6Bxwza6T36rNdvfI2pKXdcbrrFjt+EoUD6K6/5hcf/jdhK3XZm/+sfFVsG3g1120X1BLAwQUAAAACAAxQtJcmVycIxAGAACc"
    "JwAAEwAAAHhsL3RoZW1lL3RoZW1lMS54bWztWltz2jgUfu+v0Hhn9m0LxjaBtrQTc2l227SZhO1OH4URWI1seWSRhH+/RzYQ"
    "y5YN7ZJNups8BCzp+85FR+foOHnz7i5i6IaIlPJ4YNkv29a7ty/e4FcyJBFBMBmnr/DACqVMXrVaaQDDOH3JExLD3IKLCEt4"
    "FMvWXOBbGi8j1uq0291WhGlsoRhHZGB9XixoQNBUUVpvXyC05R8z+BXLVI1lowETV0EmuYi08vlsxfza3j5lz+k6HTKBbjAb"
    "WCB/zm+n5E5aiOFUwsTAamc/VmvH0dJIgILJfZQFukn2o9MVCDINOzqdWM52fPbE7Z+Mytp0NG0a4OPxeDi2y9KLcBwE4FG7"
    "nsKd9Gy/pEEJtKNp0GTY9tqukaaqjVNP0/d93+ubaJwKjVtP02t33dOOicat0HgNvvFPh8Ouicar0HTraSYn/a5rpOkWaEJG"
    "4+t6EhW15UDTIABYcHbWzNIDll4p+nWUGtkdu91BXPBY7jmJEf7GxQTWadIZljRGcp2QBQ4AN8TRTFB8r0G2iuDCktJckNbP"
    "KbVQGgiayIH1R4Ihxdyv/fWXu8mkM3qdfTrOa5R/aasBp+27m8+T/HPo5J+nk9dNQs5wvCwJ8fsjW2GHJ247E3I6HGdCfM/2"
    "9pGlJTLP7/kK6048Zx9WlrBdz8/knoxyI7vd9lh99k9HbiPXqcCzIteURiRFn8gtuuQROLVJDTITPwidhphqUBwCpAkxlqGG"
    "+LTGrBHgE323vgjI342I96tvmj1XoVhJ2oT4EEYa4pxz5nPRbPsHpUbR9lW83KOXWBUBlxjfNKo1LMXWeJXA8a2cPB0TEs2U"
    "CwZBhpckJhKpOX5NSBP+K6Xa/pzTQPCULyT6SpGPabMjp3QmzegzGsFGrxt1h2jSPHr+BfmcNQockRsdAmcbs0YhhGm78B6v"
    "JI6arcIRK0I+Yhk2GnK1FoG2camEYFoSxtF4TtK0EfxZrDWTPmDI7M2Rdc7WkQ4Rkl43Qj5izouQEb8ehjhKmu2icVgE/Z5e"
    "w0nB6ILLZv24fobVM2wsjvdH1BdK5A8mpz/pMjQHo5pZCb2EVmqfqoc0PqgeMgoF8bkePuV6eAo3lsa8UK6CewH/0do3wqv4"
    "gsA5fy59z6XvufQ9odK3NyN9Z8HTi1veRm5bxPuuMdrXNC4oY1dyzcjHVK+TKdg5n8Ds/Wg+nvHt+tkkhK+aWS0jFpBLgbNB"
    "JLj8i8rwKsQJ6GRbJQnLVNNlN4oSnkIbbulT9UqV1+WvuSi4PFvk6a+hdD4sz/k8X+e0zQszQ7dyS+q2lL61JjhK9LHMcE4e"
    "yww7ZzySHbZ3oB01+/ZdduQjpTBTl0O4GkK+A226ndw6OJ6YkbkK01KQb8P56cV4GuI52QS5fZhXbefY0dH758FRsKPvPJYd"
    "x4jyoiHuoYaYz8NDh3l7X5hnlcZQNBRtbKwkLEa3YLjX8SwU4GRgLaAHg69RAvJSVWAxW8YDK5CifEyMRehw55dcX+PRkuPb"
    "pmW1bq8pdxltIlI5wmmYE2eryt5lscFVHc9VW/Kwvmo9tBVOz/5ZrcifDBFOFgsSSGOUF6ZKovMZU77nK0nEVTi/RTO2EpcY"
    "vOPmx3FOU7gSdrYPAjK5uzmpemUxZ6by3y0MCSxbiFkS4k1d7dXnm5yueiJ2+pd3wWDy/XDJRw/lO+df9F1Drn723eP6bpM7"
    "SEycecURAXRFAiOVHAYWFzLkUO6SkAYTAc2UyUTwAoJkphyAmPoLvfIMuSkVzq0+OX9FLIOGTl7SJRIUirAMBSEXcuPv75Nq"
    "d4zX+iyBbYRUMmTVF8pDicE9M3JD2FQl867aJguF2+JUzbsaviZgS8N6bp0tJ//bXtQ9tBc9RvOjmeAes4dzm3q4wkWs/1jW"
    "Hvky3zlw2zreA17mEyxDpH7BfYqKgBGrYr66r0/5JZw7tHvxgSCb/NbbpPbd4Ax81KtapWQrET9LB3wfkgZjjFv0NF+PFGKt"
    "prGtxtoxDHmAWPMMoWY434dFmhoz1YusOY0Kb0HVQOU/29QNaPYNNByRBV4xmbY2o+ROCjzc/u8NsMLEjuHti78BUEsDBBQA"
    "AAAIADFC0lxUc8ItGQQAAOgPAAAYAAAAeGwvd29ya3NoZWV0cy9zaGVldDEueG1sjZddc5s4FIb/iob7LRaJ89GxPZO4TZ1d"
    "J/XabfdagYOtiZBYSazr/fV7BAS8U3zim4QP6fWDJB50JntjX90OwLOfhdJuGu28Lz/GsUt3UAj3wZSg8U5ubCE8ntpt7EoL"
    "Iqs7FSpORqOruBBSR7NJfW1lZxNTeSU1rCxzVVEIe7gHZfbTiEdvF9Zyu/PhQjyblGILG/Dfy5XFs7hLyWQB2kmjmYV8Gt3x"
    "j/fJRehQt/ghYe+Ojll4lBdjXsPJYzaNRlGI1sAOm1LJ+seYN+UScj8HpTAwiZhIvfwHVthsGr0Y700R7iOmFx4v5db8C7r+"
    "TVCAbRGm/KVxE9KGhmf8uwWOuucJUMfHb+QP9cDiQL0IB3Oj/pKZ302jm4hlkItK+bXZL6AdrHHIS41y9V+2b9pejCOWVg5p"
    "2s5IUEjd/Bc/20E+6sBvTnRI2g5Jzd38UE35SXgxm1izZza0xrRwUD9qM6zTSOow4xtv8a7Efn42NwUO1YE9iwImscfEcD1O"
    "2973dO+7NDWV9uy5Kl7A/r9/jCQdTtLhJHVgcjKwAPbNikzqLVv6bAjpnYT5/Dc+Go0ImIsO5oKMugcv2MaoKqwox5aykB4G"
    "keicFokTSJcd0iUZNRc5sGcj7RAF3bWlSAiKcUcxJqM+gcKRWRgVpskNodD9W5QLAuWqQ7kioz6XTip83x91hi+LleDwMDW2"
    "NFacmCw6sGW7JNiuO7ZrMupBqBTRhM5wFYUFNLye6ZCWZ0zw3HQ8N2TUFys0rmC2NFvpvEwHZ46OaGmuCJrbjuaWjFoI+2Iq"
    "y+rPwsJ4UOQrRoe1XNcEFx/1ShyRYY/W6FQJnDVIKyv9oZfAoCLptBbthkI7sjUnw36vShwgy54gk4J9saYqB5HolBbplkLq"
    "jc1p4f6BDsglfoEsY2s0plSMmsd30ho0Tumb9/7mtHiX4ZO8M5UDNsfJww81sg5CnSNwTgmc9wbntIefwEqcPI2LKvd7YYEc"
    "rXOczimn817qnLbys7F+t5doq896CxqQ8/RH+J2sFowyPO8Vz2klfxWvOGRbYCuLO13rUfKDSOeInVNi573ZOW3lldRapArY"
    "g9RCu8PJYTpH7pySO+/tzmk3/1kJ7auC3WmhDqf0/k5GC0T5nfeC57SU17jftk5mOErGZMM452idU1pPeq0ntIgXIJTfNdJk"
    "G6xwpGer5Xxwi3mO0jml9KRXekLL+BuubcA99JPQVY5FSlW/du2+fBDuHLlzSu7J0Xac1vF3bUGFjRRb4zbGFOy+ctjEndzM"
    "vJPXwCWU3pNe7wmt5Y3BYuGrD1+etzIG936DVOf4PRn0e3xUW4Ua+EnYrcTNnMKiEsvXD9coQtuUfs0JVq91wdTUnvXhDktx"
    "sKEB3s8N7nXak1DBdcX97D9QSwMEFAAAAAgAMULSXIe9kdO4AgAAvQsAAA0AAAB4bC9zdHlsZXMueG1s3VbbjpswEP0VxAeU"
    "JKQ0VBBpi7RSpbZaafehr04wYMkXaswq2a+vBxPIxbO7rfpUUIQ9x+fMxWNI1pkjp48NpSY4CC67PGyMaT9HUbdvqCDdB9VS"
    "aZFKaUGMneo66lpNSdkBSfBotVgkkSBMhttM9uJemC7Yq16aPFyE0TarlJwtcegMdikRNHgmPA8LwtlOs2EtEYwfnXkFhr3i"
    "SgfGhkLzcAmW7sXBSzeDKEcdwaTSYIych2s/d5oRDvhuVJgd6Hpno13cD9etl7cEp9WLs9XDo7Msxvll/tawzVpiDNXy3k4G"
    "zmC8gYJx/HRsbQFqTY7L1cfw3YROcVaCy7o4z3O9/rQq1oPMGXUSHR428p3SJdVT7MvwZNpmnFbG0jWrG3ga1UJdlTFK2EHJ"
    "SK0kGRI7McaBld1Tzh+h7X5WF9qHKnD987WE1gmgfqehDWgcOhk3Af1zNad9Jhv/lWzQsmdlvvQ2GznMf/XK0AdNK3YY5odq"
    "8o+pL2f11ZU6aVt+vOOsloK63N/tcJuREy9olGYv1hs03t4aqOv9Q4UHtfonKUdjkc928mIfJ2sARyUPf8Crg88Swa5n3DA5"
    "zhpWllTebKeVN2Rn300X+nZ9SSvSc/M0gXk4j7/TkvUinVY9QFrjqnn8Ddp2mUzn3/pisqQHWhbj1B6UixPjLiBcI/M74xbB"
    "OA7zI4BhfrAIMI5jYX7+p3w2aD4Ow2LbeJENytmgHMfyIcVwY378nNRe/kzTNI6TBKtoUXgjKLC6JQn8/GpYbMDA/ICnP6s1"
    "vtt4h7zeB9ievtYhWKZ4J2KZ4rUGxF83YKSpf7cxP8DAdgHrHfDv9wM95efEMewqFht2gnEkTTEEetHfo0mCVCeB278/2CmJ"
    "4zT1I4D5I4hjDIHTiCNYBBADhsTx8B28+h5Fp+9UNP9h3/4GUEsDBBQAAAAIADFC0lyXirscwAAAABMCAAALAAAAX3JlbHMv"
    "LnJlbHOdkrluwzAMQH/F0J4wB9AhiDNl8RYE+QFWog/YEgWKRZ2/r9qlcZALGXk9PBLcHmlA7TiktoupGP0QUmla1bgBSLYl"
    "j2nOkUKu1CweNYfSQETbY0OwWiw+QC4ZZre9ZBanc6RXiFzXnaU92y9PQW+ArzpMcUJpSEszDvDN0n8y9/MMNUXlSiOVWxp4"
    "0+X+duBJ0aEiWBaaRcnToh2lfx3H9pDT6a9jIrR6W+j5cWhUCo7cYyWMcWK0/jWCyQ/sfgBQSwMEFAAAAAgAMULSXDZpQ0A4"
    "AQAAKAIAAA8AAAB4bC93b3JrYm9vay54bWyNUUFuwjAQ/IrlBzQBtUhFhEtRWyTaooK4O8mGrLC9kb2Bltd3kygqUi892TO7"
    "Gs+MFxcKp5zopL6c9THTNXMzT5JY1OBMvKMGvEwqCs6wwHBMYhPAlLEGYGeTaZrOEmfQ6+Vi1NqG5BYQQ8FIXsiOOCBc4u+8"
    "g+qMEXO0yN+Z7u8WtHLo0eEVykynWsWaLq8U8Eqejd0VgazN9GQYHCAwFn/oXWdyb/LYM2zyTyNGMj1LRbDCELnf6PWNeDyD"
    "LA+oZXpGyxBWhuElUNugP3YykiK5idH3MJ5DifPwnxqpqrCAFRWtA89DjwFsZ9DHGpuolTcOMr3ByGpP6s00XSp5Zl0OCVms"
    "3fQV5iiDsC4Hk6OzEir0UL6LWBReWiq2QXVHrzO9f5g8ShuttU/CffgNmXIMOn7S8gdQSwMEFAAAAAgAMULSXCQem6KtAAAA"
    "+AEAABoAAAB4bC9fcmVscy93b3JrYm9vay54bWwucmVsc7WRPQ6DMAyFrxLlADVQqUMFTF1YKy4QBfMjEhLFrgq3L4UBkDp0"
    "YbKeLX/vyU6faBR3bqC28yRGawbKZMvs7wCkW7SKLs7jME9qF6ziWYYGvNK9ahCSKLpB2DNknu6Zopw8/kN0dd1pfDj9sjjw"
    "DzC8XeipRWQpShUa5EzCaLY2wVLiy0yWoqgyGYoqlnBaIOLJIG1pVn2wT06053kXN/dFrs3jCa7fDHB4dP4BUEsDBBQAAAAI"
    "ADFC0lxlkHmSGQEAAM8DAAATAAAAW0NvbnRlbnRfVHlwZXNdLnhtbK2TTU7DMBCFrxJlWyUuLFigphtgC11wAWNPGqv+k2da"
    "0tszTtpKoBIVhU2seN68z56XrN6PEbDonfXYlB1RfBQCVQdOYh0ieK60ITlJ/Jq2Ikq1k1sQ98vlg1DBE3iqKHuU69UztHJv"
    "qXjpeRtN8E2ZwGJZPI3CzGpKGaM1ShLXxcHrH5TqRKi5c9BgZyIuWFCKq4Rc+R1w6ns7QEpGQ7GRiV6lY5XorUA6WsB62uLK"
    "GUPbGgU6qL3jlhpjAqmxAyBn69F0MU0mnjCMz7vZ/MFmCsjKTQoRObEEf8edI8ndVWQjSGSmr3ghsvXs+0FOW4O+kc3j/Qxp"
    "N+SBYljmz/h7xhf/G87xEcLuvz+xvNZOGn/mi+E/Xn8BUEsBAhQDFAAAAAgAMULSXEbHTUiVAAAAzQAAABAAAAAAAAAAAAAA"
    "AIABAAAAAGRvY1Byb3BzL2FwcC54bWxQSwECFAMUAAAACAAxQtJcI8EtcO8AAAArAgAAEQAAAAAAAAAAAAAAgAHDAAAAZG9j"
    "UHJvcHMvY29yZS54bWxQSwECFAMUAAAACAAxQtJcmVycIxAGAACcJwAAEwAAAAAAAAAAAAAAgAHhAQAAeGwvdGhlbWUvdGhl"
    "bWUxLnhtbFBLAQIUAxQAAAAIADFC0lxUc8ItGQQAAOgPAAAYAAAAAAAAAAAAAACAgSIIAAB4bC93b3Jrc2hlZXRzL3NoZWV0"
    "MS54bWxQSwECFAMUAAAACAAxQtJch72R07gCAAC9CwAADQAAAAAAAAAAAAAAgAFxDAAAeGwvc3R5bGVzLnhtbFBLAQIUAxQA"
    "AAAIADFC0lyXirscwAAAABMCAAALAAAAAAAAAAAAAACAAVQPAABfcmVscy8ucmVsc1BLAQIUAxQAAAAIADFC0lw2aUNAOAEA"
    "ACgCAAAPAAAAAAAAAAAAAACAAT0QAAB4bC93b3JrYm9vay54bWxQSwECFAMUAAAACAAxQtJcJB6boq0AAAD4AQAAGgAAAAAA"
    "AAAAAAAAgAGiEQAAeGwvX3JlbHMvd29ya2Jvb2sueG1sLnJlbHNQSwECFAMUAAAACAAxQtJcZZB5khkBAADPAwAAEwAAAAAA"
    "AAAAAAAAgAGHEgAAW0NvbnRlbnRfVHlwZXNdLnhtbFBLBQYAAAAACQAJAD4CAADREwAAAAA="
)



def _sample_bytes(b64_str: str) -> bytes:
    return base64.b64decode(b64_str)


class _FakeUpload:
    """Mimics the subset of Streamlit's UploadedFile interface this app uses,
    so embedded sample bytes can be fed through the exact same code path as
    a real file uploaded via st.file_uploader."""

    def __init__(self, name: str, data: bytes):
        self.name = name
        self._data = data

    def getvalue(self) -> bytes:
        return self._data


SAMPLES = {
    "one_file_two_tabs": {
        "label": "1 file, 2 tabs",
        "description": "A single workbook containing both lists, on separate tabs.",
        "filename": "Sample - 1 file with 2 tabs.xlsx",
        "bytes": _sample_bytes(SAMPLE_1FILE_2TABS_B64),
    },
    "two_files_reference": {
        "label": "2 files — Reference List",
        "description": "The reference/standard list, as its own workbook.",
        "filename": "Sample - Reference List.xlsx",
        "bytes": _sample_bytes(SAMPLE_REFERENCE_LIST_B64),
    },
    "two_files_to_map": {
        "label": "2 files — List To Map",
        "description": "The list to be matched against the reference, as its own workbook.",
        "filename": "Sample - List To Map.xlsx",
        "bytes": _sample_bytes(SAMPLE_LIST_TO_MAP_B64),
    },
}


# ============================================================================
# 1. NORMALISE COMPANY NAMES
# ============================================================================
def clean(name: str) -> str:
    """Normalise a company name: lowercase, strip accents, strip legal/status suffixes, strip punctuation."""
    if not isinstance(name, str):
        return ""
    s = name.lower()
    s = unicodedata.normalize("NFKD", s)
    s = "".join(c for c in s if not unicodedata.combining(c))
    s = re.sub(
        r"\b(limited|ltd|llp|llc|plc|inc|incorporated|gmbh|srl|bv|nv|sa"
        r"|lp|corp|corporation|company|ag|kg|oy|ab|pty|pte"
        r"|in administration|in liquidation|in receivership|dissolved)\b",
        " ", s
    )
    s = re.sub(r"\bt\/as?\b.*$", " ", s)
    s = re.sub(r"[^a-z0-9\s]", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def combined_score(a_clean: str, b_clean: str) -> float:
    """Blend token_sort_ratio (handles reordered words) with ratio (penalises length mismatches)."""
    s1 = fuzz.token_sort_ratio(a_clean, b_clean)
    s2 = fuzz.ratio(a_clean, b_clean)
    return 0.7 * s1 + 0.3 * s2


# ============================================================================
# 2. DETECT COMPANY-NAME COLUMN & EXTRACT NAMES
# ============================================================================
def find_company_column(df: pd.DataFrame) -> int:
    headers = [str(c).strip().lower() for c in df.columns]
    for i, h in enumerate(headers):
        if "company" in h and "name" in h:
            return i
    for i, h in enumerate(headers):
        if "name" in h or "company" in h:
            return i
    for i, col in enumerate(df.columns):
        non_null = df[col].dropna()
        if len(non_null) > 0 and non_null.astype(str).str.match(r"^[A-Za-z]").mean() > 0.5:
            return i
    return 0


def extract_names(df_raw: pd.DataFrame) -> list[str]:
    df = df_raw.dropna(how="all")
    if df.empty:
        return []
    col_idx = find_company_column(df)
    col = df.iloc[:, col_idx].dropna().astype(str).str.strip()
    return [c for c in col.tolist() if c and c.lower() not in ("name", "company name", "nan")]


# ============================================================================
# 3. MATCHING
# ============================================================================
def match_all(reference_names: list[str], other_names: list[str], threshold: float):
    other_cleaned = [clean(n) for n in other_names]
    valid_idx = [i for i, c in enumerate(other_cleaned) if c]
    valid_other_clean = [other_cleaned[i] for i in valid_idx]
    valid_other_orig = [other_names[i] for i in valid_idx]

    results = []
    for ref_orig in reference_names:
        ref_clean = clean(ref_orig)
        if not ref_clean or not valid_other_clean:
            results.append((ref_orig, "", 0.0))
            continue

        coarse = process.extract(
            ref_clean, valid_other_clean, scorer=fuzz.token_sort_ratio, limit=5
        )
        best_score, best_name = 0.0, ""
        for _, coarse_score, idx in coarse:
            if coarse_score < threshold - 8:
                continue
            sc = combined_score(ref_clean, valid_other_clean[idx])
            if sc > best_score:
                best_score, best_name = sc, valid_other_orig[idx]

        results.append((ref_orig, best_name if best_score >= threshold else "", round(best_score, 1)))

    return results


# ============================================================================
# 4. BUILD OUTPUT EXCEL FILE (returns bytes for Streamlit download)
# ============================================================================
HEADER_FILL = PatternFill("solid", fgColor="4472C4")
HEADER_FONT = Font(bold=True, color="FFFFFF", name="Arial", size=11)
ALT_FILL = PatternFill("solid", fgColor="DCE6F1")
DATA_FONT = Font(name="Arial", size=10)


def write_header(ws, headers):
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center")


def autosize_columns(ws, widths):
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = min(max(width, 12), 70)


def write_result_sheet(wb, results, reference_label, other_label):
    ws = wb.create_sheet("Result")
    write_header(ws, [reference_label, other_label, "Match score"])
    for row_idx, (ref, match, score) in enumerate(results, start=2):
        fill = ALT_FILL if row_idx % 2 == 0 else None
        for col, val in enumerate([ref, match, score if match else ""], 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.font = DATA_FONT
            if fill:
                cell.fill = fill
    autosize_columns(ws, [55, 55, 14])
    ws.freeze_panes = "A2"


def write_source_sheet(wb, df_raw, sheet_name):
    ws = wb.create_sheet(sheet_name[:31])
    headers = [str(c) for c in df_raw.columns]
    write_header(ws, headers)
    for row_idx, (_, row) in enumerate(df_raw.iterrows(), start=2):
        fill = ALT_FILL if row_idx % 2 == 0 else None
        for col, val in enumerate(row.tolist(), 1):
            cell = ws.cell(row=row_idx, column=col, value=None if pd.isna(val) else val)
            cell.font = DATA_FONT
            if fill:
                cell.fill = fill
    autosize_columns(ws, [max(18, len(h) + 2) for h in headers])
    ws.freeze_panes = "A2"


def build_output_excel(results, reference_label, other_label, reference_raw, other_raw) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)
    write_result_sheet(wb, results, reference_label, other_label)
    write_source_sheet(wb, reference_raw, f"Source - {reference_label}")
    write_source_sheet(wb, other_raw, f"Source - {other_label}")
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ============================================================================
# 5. READ UPLOADED FILE → DICT {sheet_name: df}
# ============================================================================
@st.cache_data(show_spinner=False)
def read_excel_sheets(file_bytes: bytes) -> dict[str, pd.DataFrame]:
    return pd.read_excel(io.BytesIO(file_bytes), sheet_name=None, dtype=str)


# ============================================================================
# SAMPLE DATA UI — download buttons + "run this sample" buttons
# ============================================================================
def render_sample_picker():
    with st.expander("📂 Don't have files handy? Try the built-in sample data", expanded=False):
        st.write(
            "Two ways to try the app: download a sample to see the expected format, "
            "or click **Run this sample** to load it straight in below."
        )
        tab1, tab2 = st.tabs(["Sample: 1 file, 2 tabs", "Sample: 2 separate files"])

        with tab1:
            s = SAMPLES["one_file_two_tabs"]
            st.caption(s["description"])
            c1, c2 = st.columns(2)
            with c1:
                st.download_button(
                    "⬇️ Download sample file",
                    data=s["bytes"],
                    file_name=s["filename"],
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="dl_one_file",
                )
            with c2:
                if st.button("▶️ Run this sample", key="run_one_file", type="primary"):
                    st.session_state["sample_mode"] = "1 Excel file (multiple tabs)"
                    st.session_state["sample_single_file"] = s
                    st.session_state.pop("sample_indue_file", None)
                    st.session_state.pop("sample_other_file", None)
                    st.rerun()

        with tab2:
            ref = SAMPLES["two_files_reference"]
            other = SAMPLES["two_files_to_map"]
            st.caption(f"{ref['description']} Paired with: {other['description'].lower()}")
            c1, c2, c3 = st.columns(3)
            with c1:
                st.download_button(
                    "⬇️ Reference List",
                    data=ref["bytes"],
                    file_name=ref["filename"],
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="dl_ref",
                )
            with c2:
                st.download_button(
                    "⬇️ List To Map",
                    data=other["bytes"],
                    file_name=other["filename"],
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="dl_other",
                )
            with c3:
                if st.button("▶️ Run this sample", key="run_two_files", type="primary"):
                    st.session_state["sample_mode"] = "2 separate Excel files"
                    st.session_state["sample_indue_file"] = ref
                    st.session_state["sample_other_file"] = other
                    st.session_state.pop("sample_single_file", None)
                    st.rerun()


# ============================================================================
# UI
# ============================================================================
st.title("🔎 Company Name Matcher")
st.write(
    "Match company names between a **Reference list** (your standard/master list) and "
    "**another list** using fuzzy matching. Supports uploading **1 Excel file with multiple "
    "tabs** or **2 separate Excel files**."
)

render_sample_picker()

mode_options = ["1 Excel file (multiple tabs)", "2 separate Excel files"]
default_mode = st.session_state.get("sample_mode", mode_options[0])
mode = st.radio(
    "Choose how to upload your data:",
    mode_options,
    index=mode_options.index(default_mode),
    horizontal=True,
)

threshold = st.slider(
    "Similarity threshold to accept a match (0–100)",
    min_value=50, max_value=100, value=86, step=1,
    help="Scores below this threshold leave the match cell blank (treated as no match found)."
)

reference_raw, other_raw, reference_label, other_label = None, None, None, None

if mode == "1 Excel file (multiple tabs)":
    sample_single = st.session_state.get("sample_single_file")
    uploaded = st.file_uploader("Upload Excel file (.xlsx)", type=["xlsx", "xlsm"])

    if uploaded:
        st.session_state.pop("sample_single_file", None)
        sample_single = None
        source = uploaded
    elif sample_single:
        st.info(f"Using sample file **{sample_single['filename']}** — upload your own file above to replace it.")
        source = _FakeUpload(sample_single["filename"], sample_single["bytes"])
    else:
        source = None

    if source:
        sheets = read_excel_sheets(source.getvalue())
        sheet_names = list(sheets.keys())
        if len(sheet_names) < 2:
            st.error("This file only has 1 tab. At least 2 tabs are needed to compare.")
        else:
            col1, col2 = st.columns(2)
            with col1:
                reference_label = st.selectbox("Reference list (tab)", sheet_names, index=0)
            with col2:
                remaining = [n for n in sheet_names if n != reference_label]
                other_label = st.selectbox("List to map (tab)", remaining, index=0)
            reference_raw = sheets[reference_label]
            other_raw = sheets[other_label]

else:
    sample_ref = st.session_state.get("sample_indue_file")
    sample_other = st.session_state.get("sample_other_file")

    col1, col2 = st.columns(2)
    with col1:
        reference_file = st.file_uploader("Upload Reference list file", type=["xlsx", "xlsm"], key="indue")
    with col2:
        other_file = st.file_uploader("Upload the list to map", type=["xlsx", "xlsm"], key="other")

    if reference_file:
        st.session_state.pop("sample_indue_file", None)
        sample_ref = None
    if other_file:
        st.session_state.pop("sample_other_file", None)
        sample_other = None

    ref_source = reference_file or (_FakeUpload(sample_ref["filename"], sample_ref["bytes"]) if sample_ref else None)
    other_source = other_file or (_FakeUpload(sample_other["filename"], sample_other["bytes"]) if sample_other else None)

    if sample_ref and not reference_file:
        st.caption(f"Using sample: **{sample_ref['filename']}** — upload your own file above to replace it.")
    if sample_other and not other_file:
        st.caption(f"Using sample: **{sample_other['filename']}** — upload your own file above to replace it.")

    if ref_source and other_source:
        reference_sheets = read_excel_sheets(ref_source.getvalue())
        other_sheets = read_excel_sheets(other_source.getvalue())

        def pick_sheet(sheets):
            names = list(sheets.keys())
            return names[0], sheets[names[0]]

        _, reference_raw = pick_sheet(reference_sheets)
        _, other_raw = pick_sheet(other_sheets)
        reference_label = ref_source.name.rsplit(".", 1)[0]
        other_label = other_source.name.rsplit(".", 1)[0]

# ============================================================================
# RUN MATCHING
# ============================================================================
if reference_raw is not None and other_raw is not None:
    reference_names = extract_names(reference_raw)
    other_names = extract_names(other_raw)

    st.write(f"📋 **{reference_label}**: {len(reference_names)} companies &nbsp;|&nbsp; **{other_label}**: {len(other_names)} companies")

    if st.button("🚀 Run matching", type="primary"):
        if not reference_names or not other_names:
            st.error("No company names found in one of the two data sources. Please check your file(s).")
        else:
            with st.spinner("Matching..."):
                results = match_all(reference_names, other_names, threshold)

            matched = sum(1 for _, k, _ in results if k)
            st.success(f"Done! Matched {matched} / {len(reference_names)} reference companies.")

            result_df = pd.DataFrame(results, columns=[reference_label, other_label, "Match score"])
            result_df["Match score"] = result_df.apply(
                lambda r: r["Match score"] if r[other_label] else None, axis=1
            )
            st.dataframe(result_df, use_container_width=True, height=400)

            output_bytes = build_output_excel(results, reference_label, other_label, reference_raw, other_raw)
            st.download_button(
                label="⬇️ Download result file (.xlsx)",
                data=output_bytes,
                file_name="Matching result - OUTPUT.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

            with st.expander("View original data"):
                tab1, tab2 = st.tabs([f"Source - {reference_label}", f"Source - {other_label}"])
                with tab1:
                    st.dataframe(reference_raw, use_container_width=True)
                with tab2:
                    st.dataframe(other_raw, use_container_width=True)
else:
    st.info("👆 Upload a file, or run a sample above, to get started.")
